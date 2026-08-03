"""
📊 Format Laporan Komprehensif Biodiversitas
============================================
Menghasilkan tabulasi format profesional dalam SATU SHEET dengan:
- Baris: Kategori data + satuan (dinamis sesuai mode tampilan)
- Kolom: Stasiun (ID) + Rata-rata
- Support mode: "Data Asli" atau "Data Terkonversi"
"""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Import konstanta konversi kelimpahan
from konversi import KONVERSI_KELIMPAHAN_PER_HA


def get_satuan_labels(mode_tampilan="Data Terkonversi"):
    """
    Dapatkan label satuan sesuai mode tampilan.
    
    Parameters:
    -----------
    mode_tampilan : str
        "Data Asli" atau "Data Terkonversi"
    
    Returns:
    --------
    dict
        {
            'kelimpahan': 'ind/ha' or 'ind/stasiun',
            'biomassa': 'kg/ha' or 'g/stasiun',
        }
    """
    if mode_tampilan == "Data Terkonversi":
        return {
            'kelimpahan': 'ind/ha',
            'biomassa': 'kg/ha',
        }
    else:  # Data Asli
        return {
            'kelimpahan': 'ind/stasiun',
            'biomassa': 'g/stasiun',
        }


def prepare_comprehensive_report(df_merge, df_indeks, mode_tampilan="Data Terkonversi", stasiun_order=None):
    """
    Siapkan laporan komprehensif dalam FORMAT TUNGGAL dengan struktur:
    - Baris 1: Header (Kategori | Stasiun1 | Stasiun2 | ... | Rata-rata)
    - Baris 2+: Kategori data (Indeks, Kelimpahan, Biomassa, Keanekaragaman)
    
    Parameters:
    -----------
    df_merge : DataFrame
        Data hasil merge dari struktur.py (dengan Biomassa, Kelimpahan, Keanekaragaman)
    df_indeks : DataFrame
        Data indeks per stasiun (Shannon, Simpson, Evenness)
    mode_tampilan : str
        "Data Asli" atau "Data Terkonversi" (default: "Data Terkonversi")
    stasiun_order : list, optional
        Urutan stasiun yang diinginkan
    
    Returns:
    --------
    DataFrame
        Tabulasi lengkap dengan semua metrik
    """
    
    if stasiun_order is None:
        stasiun_order = sorted(df_indeks["Stasiun"].unique())
    
    # Pastikan kolom Stasiun ada di df_merge
    if "Stasiun" not in df_merge.columns:
        raise ValueError("df_merge harus memiliki kolom 'Stasiun'")
    
    # Dapatkan label satuan sesuai mode
    satuan = get_satuan_labels(mode_tampilan)
    
    # ========== BUAT TABULASI TUNGGAL ==========
    all_rows = []
    
    # ========== 1. INDEKS EKOLOGI ==========
    df_indeks_report = df_indeks[["Stasiun", "Shannon", "Simpson", "Evenness"]].copy()
    df_indeks_report = df_indeks_report[df_indeks_report["Stasiun"].isin(stasiun_order)]
    df_indeks_report = df_indeks_report.set_index("Stasiun").reindex(stasiun_order)
    
    # Bagian Indeks header
    all_rows.append({
        'Kategori': '=== INDEKS EKOLOGI ===',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    for idx_name in ["Shannon", "Simpson", "Evenness"]:
        row_data = {"Kategori": f"Indeks {idx_name}"}
        values = []
        
        for stasiun in stasiun_order:
            value = df_indeks_report.loc[stasiun, idx_name] if stasiun in df_indeks_report.index else 0
            row_data[stasiun] = round(value, 4)
            values.append(value)
        
        rata_rata = sum(values) / len(values) if values else 0
        row_data["Rata-rata"] = round(rata_rata, 4)
        all_rows.append(row_data)
    
    # Blank row
    all_rows.append({
        'Kategori': '',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    # ========== 2. KELIMPAHAN PER KELOMPOK ==========
    kelompok_list = sorted(df_merge["Kelompok"].unique())
    
    all_rows.append({
        'Kategori': '=== KELIMPAHAN (Individu) ===',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    for kelompok in kelompok_list:
        df_k = df_merge[df_merge["Kelompok"] == kelompok]
        row_data = {"Kategori": f"Kelimpahan {kelompok} ({satuan['kelimpahan']})"}
        values = []
        
        for stasiun in stasiun_order:
            df_st = df_k[df_k["Stasiun"] == stasiun]
            # Gunakan kolom 'Kelimpahan' jika tersedia (lebih akurat), fallback ke hitung baris
            if "Kelimpahan" in df_st.columns:
                # Kelimpahan di df_merge adalah per spesies; jumlahkan untuk mendapat total individu per stasiun
                kelimpahan = df_st["Kelimpahan"].sum()
            else:
                kelimpahan = len(df_st)
            # Terapkan konversi bila mode tampilan meminta Data Terkonversi
            if mode_tampilan == "Data Terkonversi":
                kelimpahan_display = kelimpahan * KONVERSI_KELIMPAHAN_PER_HA
                row_data[stasiun] = round(kelimpahan_display, 2)
                values.append(kelimpahan_display)
            else:
                # Data asli: tampilkan sebagai integer (jumlah individu dalam 350 m²)
                row_data[stasiun] = int(kelimpahan)
                values.append(kelimpahan)
        
        rata_rata = sum(values) / len(stasiun_order) if values else 0
        row_data["Rata-rata"] = round(rata_rata, 2) if mode_tampilan == "Data Terkonversi" else int(round(rata_rata))
        all_rows.append(row_data)
    
    # Blank row
    all_rows.append({
        'Kategori': '',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    # ========== 3. BIOMASSA PER KELOMPOK ==========
    all_rows.append({
        'Kategori': '=== BIOMASSA ===',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    for kelompok in kelompok_list:
        df_k = df_merge[df_merge["Kelompok"] == kelompok]
        row_data = {"Kategori": f"Biomassa {kelompok} ({satuan['biomassa']})"}
        values = []
        
        for stasiun in stasiun_order:
            df_st = df_k[df_k["Stasiun"] == stasiun]
            biomassa = df_st["Biomassa"].sum() if "Biomassa" in df_st.columns else 0
            row_data[stasiun] = round(biomassa, 2)
            values.append(biomassa)
        
        rata_rata = sum(values) / len(stasiun_order) if values else 0
        row_data["Rata-rata"] = round(rata_rata, 2)
        all_rows.append(row_data)
    
    # Blank row
    all_rows.append({
        'Kategori': '',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    # ========== 4. KEANEKARAGAMAN JENIS PER KELOMPOK ==========
    all_rows.append({
        'Kategori': '=== KEANEKARAGAMAN JENIS ===',
        **{stasiun: '' for stasiun in stasiun_order},
        'Rata-rata': ''
    })
    
    for kelompok in kelompok_list:
        df_k = df_merge[df_merge["Kelompok"] == kelompok]
        row_data = {"Kategori": f"Keanekaragaman Jenis {kelompok}"}
        values = []
        
        for stasiun in stasiun_order:
            df_st = df_k[df_k["Stasiun"] == stasiun]
            n_spesies = df_st["Spesies"].nunique() if "Spesies" in df_st.columns else 0
            row_data[stasiun] = n_spesies
            values.append(n_spesies)
        
        rata_rata = sum(values) / len(stasiun_order) if values else 0
        row_data["Rata-rata"] = round(rata_rata, 1)
        all_rows.append(row_data)
    
    # ========== BUAT DATAFRAME FINAL ==========
    df_report = pd.DataFrame(all_rows)
    
    return df_report


def generate_comprehensive_excel(df_merge, df_indeks, mode_tampilan="Data Terkonversi", stasiun_order=None):
    """
    Generate Excel SINGLE-SHEET dengan format laporan komprehensif.
    
    Parameters:
    -----------
    df_merge : DataFrame
    df_indeks : DataFrame
    mode_tampilan : str
        "Data Asli" atau "Data Terkonversi" (default: "Data Terkonversi")
    stasiun_order : list, optional
    
    Returns:
    --------
    BytesIO
        Excel file buffer siap untuk download
    """
    
    df_report = prepare_comprehensive_report(df_merge, df_indeks, mode_tampilan, stasiun_order)
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Hanya 1 sheet
        df_report.to_excel(writer, sheet_name="Laporan Komprehensif", index=False)
        
        worksheet = writer.sheets["Laporan Komprehensif"]
        
        # 🖌️ Format header
        header_fill = PatternFill(start_color="003f5c", end_color="003f5c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, col_name in enumerate(df_report.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # 🎨 Format data cells
        for row_num, row in enumerate(dataframe_to_rows(df_report, index=False, header=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Kolom pertama (Kategori)
                if col_num == 1:
                    # Section header (===) - bold, light gray background
                    if isinstance(value, str) and '===' in value:
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1f5a7a", end_color="1f5a7a", fill_type="solid")
                        cell.alignment = Alignment(horizontal='left', wrap_text=True)
                    # Blank rows
                    elif value == '':
                        cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                    # Regular kategori
                    else:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='left', wrap_text=True)
                
                # Kolom Rata-rata (terakhir) - highlight
                elif col_num == len(df_report.columns):
                    if isinstance(value, str) and '===' in value:
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1f5a7a", end_color="1f5a7a", fill_type="solid")
                    elif value == '':
                        cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='right')
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            if isinstance(value, float):
                                cell.number_format = '0.00'
                
                # Data cells (Stasiun) - right align
                else:
                    if isinstance(value, str) and '===' in value:
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1f5a7a", end_color="1f5a7a", fill_type="solid")
                    elif value == '':
                        cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                    else:
                        cell.alignment = Alignment(horizontal='right')
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            if isinstance(value, float):
                                cell.number_format = '0.00'
        
        # 📐 Auto column width
        for col_num, col_name in enumerate(df_report.columns, 1):
            col_letter = worksheet.cell(row=1, column=col_num).column_letter
            max_length = len(str(col_name))
            for row in worksheet.iter_rows(max_row=len(df_report) + 1, min_col=col_num, max_col=col_num):
                try:
                    if len(str(row[0].value)) > max_length:
                        max_length = len(str(row[0].value))
                except:
                    pass
            
            # Kategori kolom lebih lebar
            if col_num == 1:
                worksheet.column_dimensions[col_letter].width = max(max_length + 3, 30)
            else:
                worksheet.column_dimensions[col_letter].width = max(max_length + 2, 14)
        
        # Set row height untuk header
        worksheet.row_dimensions[1].height = 30
    
    buffer.seek(0)
    return buffer
